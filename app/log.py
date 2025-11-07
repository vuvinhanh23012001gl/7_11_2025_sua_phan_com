
# -*- coding: utf-8 -*-
# -----------------------------------------------------------------------------
# Title      : Check OIL bivn / Module log
# Description: Module log
# Author     : Vu Vinh Anh
# Email      : anh.vu@example.com
# Created    : 2025-06-30
# Version    : 0.1
# License    : MIT
# -----------------------------------------------------------------------------
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from obj_log import safe_put_queue,debug_print
import time
import logging
import os

class Log:
    """Lớp này lấy thông tin của log phần mềm và thực hiện bật tắt log phần mềm"""
    from folder_create import Create
    obj_folder = Create()
    characters_check = "date_"
    def __init__(self,obj_config_software=None,name="app"):
        self.log_file = None
        self.name = name
        self.console_enabled = False
        self.file_enabled = False
        self.logger = logging.getLogger(name)
        self.logger.setLevel(logging.DEBUG)  # Cho phép tất cả mức log
        self.logger.handlers.clear()  # Xóa handler cũ tránh bị log trùng
        self.formatter = logging.Formatter(
            "%(asctime)s [%(levelname)s]:%(message)s"
        )
        self.obj_config_software = obj_config_software
        if self.obj_config_software :
            self.delete_folder_old_log_software()

            open_log_console = self.obj_config_software.get_log_console()
            if open_log_console:
                self.enable_console()
            else:
                self.disable_console()

            self.log_file = self.create_file_log_software()
            if self.log_file:
                self.enable_file()
            else:
                self.disable_file()


    def get_time_software(self):
        """Lấy thời gian cho phép log phan mem được lưu nếu được bật"""
        return self.obj_config_software.GetTimeSaveLogSoftware()

    def get_open_log_software(self):
        """lấy quyền lưu log"""
        return self.obj_config_software.get_log_software()
    def get_open_log_console(self):
        return self.obj_config_software.get_log_console()
    def get_path_save_software(self):
        """Tra ve duong dan luu Fodel log log_software"""
        return self.obj_config_software.get_path_log_software()


    def get_list_file_in_folder_log_sofware(self)->list:
        """Hàm này trả về danh sách folder hiện có trong folder software"""
        return Log.obj_folder.get_list_folder_in_folder(self.get_path_save_software())

    def get_list_find_old_sofware(self,days_threshold):
        """Trả về đường danh sách tên folder có days_threshold không thỏa mãn để xóa """
        list_file = self.get_list_file_in_folder_log_sofware()
        debug_print("Danh sach file excell hiện có trong thư mục là:",list_file)
        if  list_file:
            arr_old_file  = Log.obj_folder.get_old_folders_by_threshold(Log.characters_check,list_file,days_threshold)
            debug_print(f"Danh sách file cũ hơn {days_threshold} ngày để xóa",arr_old_file)
            return arr_old_file
        else :
            debug_print("Danh sách trong folder excell rỗng")
            return None

    def delete_folder_old_log_software(self):
        arr_file_old = self.get_list_find_old_sofware(self.get_time_software())
        if arr_file_old:
            debug_print("---Bắt đầu xóa file quá hạn --")
            for file_delete in arr_file_old:
                path_file_delete = Log.obj_folder.find_file(self.get_path_save_software(),file_delete)
                debug_print("Thư mục cần xóa",path_file_delete)
                if path_file_delete:
                    Log.obj_folder.delete_folder(path_file_delete)
            debug_print("--Xóa thành công folder--")

    def create_file_log_software(self):
        """nếu cho phép bật được bật thì sẽ tạo ra folder kiểu date_ngày tạo . và tạo file text theo ngày và giờ tạo file.
        nếu open:trả thực hiện mở file và trả về đường dẫn file
        nếu k open: thì trả về None
        """
        open_log_software = self.get_open_log_software()  # đổi đúng tên hàm getter
        if not open_log_software:
            debug_print("Hiện tại đang tắt log software")
            return None
        # Lấy đường dẫn thư mục log phần mềm
        path_log_folder_software = self.obj_config_software.get_path_log_software()
        today = datetime.now().strftime("%Y-%m-%d")
        name_folder = f"date_{today}"
        path_foder = os.path.join(path_log_folder_software, name_folder)
        return Log.obj_folder.create_file_text_log(path_foder,"txt")

    def log_and_print(self, msg, value=None, level="info"):
        # Ghép message nếu có value
        full_msg = f"{msg}: {value}" if value is not None else msg
        if level == "debug":
            self.logger.debug(full_msg)
        elif level == "warning":
            self.logger.warning(full_msg)
        elif level == "error":
            self.logger.error(full_msg)
        elif level == "critical":
            self.logger.critical(full_msg)
        else:
            self.logger.info(full_msg)

    def enable_console(self):
            debug_print("Bật Log console")
            if not self.console_enabled:
                ch = logging.StreamHandler()
                ch.setLevel(logging.DEBUG)
                ch.setFormatter(self.formatter)
                self.logger.addHandler(ch)
                self.console_enabled = True

    def disable_console(self):
            debug_print("Tắt Log console")
            for h in list(self.logger.handlers):
                if isinstance(h, logging.StreamHandler):
                    self.logger.removeHandler(h)
            self.console_enabled = False

    def enable_file(self):
            debug_print("Bật Log File")
            if not self.file_enabled:
                os.makedirs(os.path.dirname(self.log_file) or ".", exist_ok=True)
                debug_print("Đường dẫn file log:", self.log_file)
                fh = logging.FileHandler(self.log_file, encoding="utf-8")
                fh.setLevel(logging.DEBUG)
                fh.setFormatter(self.formatter)
                self.logger.addHandler(fh)
                self.file_enabled = True

    def disable_file(self):
            debug_print("Tắt Log File")
            for h in list(self.logger.handlers):
                if isinstance(h, logging.FileHandler):
                    self.logger.removeHandler(h)
            self.file_enabled = False

    # ===============================
    # Các hàm log tiện dụng
    # ===============================
    def debug(self, msg):
        self.logger.debug(msg)


    def info(self, msg):
        self.logger.info(msg)

    def warning(self, msg):
        self.logger.warning(msg)

    def error(self, msg):
        self.logger.error(msg)

    def critical(self, msg):
        self.logger.critical(msg)
    def update_log_state(self):
        """Kiểm tra và cập nhật trạng thái log theo obj_config_software (real-time)."""
        if not self.obj_config_software:
            return

        # --- Console log ---
        open_console = self.obj_config_software.get_log_console()
        if open_console and not self.console_enabled:
            self.enable_console()
        elif not open_console and self.console_enabled:
            self.disable_console()

        # --- File log ---
        open_file = self.obj_config_software.get_log_software()
        if open_file:
            if not self.file_enabled:
                self.log_file = self.create_file_log_software()
                if self.log_file:
                    self.enable_file()
        else:
            if self.file_enabled:
                self.disable_file()


#==================================Hàm chạy kiểm thử====================================================#


# from config_software import OilDetectionSystem
# obj_config_software = OilDetectionSystem()
# obj_log_data = Log(obj_config_software)
# obj_log_data.info("wewewe232323232we")
# obj_log_data.info("wewewe232323232we")
# obj_log_data.info("wewewe232323232we")
# obj_log_data.info("wewewe232323232we")
# obj_log_data.info("wewewe232323232we")
# obj_log_data.info("wewewe232323232we")
# obj_log_data.info("wewewe232323232we")
# obj_log_data.info("wewewe232323232we")

# path_file  = obj_log_data.create_file_log_software()
# print(path_file)
# print(obj_log_data.get_list_find_old_sofware(1))
# obj_log_data.delete_folder_old_log_software()
#lAY THOI GIAN LUU LOG SOFTWARE
# print(obj_log_data.get_time_software())
# print(obj_log_data.get_open_log_software())
# print(obj_log_data.get_path_save_software())
# print(obj_log_data.get_list_file_in_folder_log_sofware())


class log_excell:
    '''Lớp này mỗi khi gọi sẽ tạo ra 1 file logexcell nếu được bật log excell ở trong config nếu tắt thì sẽ không tự động tạo ra 1 file'''
    from folder_create import Create
    obj_folder = Create()
    characters_check = "date_"
    def __init__(self,obj_config_software):
        debug_print("----------------------- Init File excell------------------------")
        self.wb = None
        self.ws = None

        self.obj_config_software = obj_config_software
        self.path_file_save_log_excell = self.create_file_excell()  # self.path_file_save_log_excell  se cho co the bang none neu khong duoc phep tao file
        if self.path_file_save_log_excell:
            self.delete_file_old()
            self.write_file_excel(["Thời gian","Tên người thao tác","Tên User","Loại","Nhà máy","Truyền","Trạng thái phán định","Ghi chú lỗi"])
        debug_print("-----------------------End Init File excell------------------------")
    def get_path_file_save_log_excell(self):  #ok
        """Tra ve path File luu log hien tai"""
        return self.path_file_save_log_excell
    def get_time(self): #ok
        """Lấy thời gian cho phép log được lưu nếu được bật"""
        return self.obj_config_software.GetTimeSaveLogExcell()

    def get_open_log_excell(self): #ok
        """lấy quyền lưu log"""
        return  self.obj_config_software.get_log_product()

    def get_path_folder_log_excell(self): #ok
        """Tra ve path folder luu log hien tai"""
        return self.obj_config_software.get_path_log_product()


    def get_list_folder_log_excell(self)->list: #ok
        """Hàm này trả về danh sách foder hiện có trong folder excell"""
        return log_excell.obj_folder.get_list_folder_in_folder(self.get_path_folder_log_excell())

    def create_file_excell(self): #ok
        """Tạo File Excell nếu trong config software cho phép
        input :self
        output:trả về đường dẫn file nếu cho phép tạo
        trả về None nếu không cho phép tạo
        """
        if self.get_open_log_excell():
            debug_print("Log Excell đang bật")
            path_excell =  self.get_path_folder_log_excell()
            today = datetime.now().strftime("%Y-%m-%d")
            name_folder = f"date_{today}"
            path_foder = os.path.join(path_excell,name_folder)
            file_path = log_excell.obj_folder.create_file_log(path_foder)
            debug_print("Path excell ghi dữ liệu:",file_path)
            return file_path
        else:
            debug_print("Log Excell đang tắt")
            return None

    def get_list_find_old(self,days_threshold): #ok
        """Trả về đường danh sách tên folder có days_threshold không thỏa mãn để xóa """
        list_file = self.get_list_folder_log_excell()
        if  list_file:
            arr_old_file = log_excell.obj_folder.get_old_folders_by_threshold(log_excell.characters_check,list_file,days_threshold)
            debug_print(f"Danh sách file cũ hơn {days_threshold} ngày để xóa",arr_old_file)
            return arr_old_file
        else :
            debug_print("Danh sách trong folder excell rỗng hoặc không tồn tại file nào trong đó")
            return None

    def delete_file_old(self): # oke
        arr_file_old = self.get_list_find_old(self.get_time())
        if arr_file_old:
            debug_print("---Xóa File Excell quá hạn \n Bắt đầu xóa --")
            for folder_name in arr_file_old:
                path_file_delete = log_excell.obj_folder.find_file(self.get_path_folder_log_excell(),folder_name)
                debug_print("Path Folder cần xóa",path_file_delete)
                if path_file_delete:
                    log_excell.obj_folder.delete_folder(path_file_delete)
            debug_print("--Xóa thành công file--")


    def write_file_excel(self, row: list):
        """
        Ghi 1 dòng dữ liệu vào file Excel hiện tại.
        Nếu file chưa tồn tại -> tạo mới.
        :param row: list chứa dữ liệu tương ứng 1 dòng
        """
        if not self.get_open_log_excell():
            debug_print("⚠️ Chức năng log chưa được bật, không lưu Excel")
            return None

        file_path = self.get_path_file_save_log_excell()

        # 🔹 Nếu file chưa tồn tại -> tạo mới
        if not os.path.exists(file_path):
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Log Data"
            self.wb.save(file_path)
        else:
            self.wb = load_workbook(file_path)
            self.ws = self.wb.active

        # 🔹 Ghi dữ liệu vào dòng mới
        self.ws.append(row)

        # 🔹 Tự động căn chỉnh độ rộng cột cho đẹp
        for col in self.ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 4
            self.ws.column_dimensions[col_letter].width = adjusted_width

        # 🔹 Lưu file lại
        self.wb.save(file_path)
        # print(f"✅ Đã lưu dòng dữ liệu vào: {file_path}")

        return file_path
    def update_log_state(self):
        """Cập nhật realtime trạng thái log Excel."""
        open_log_excel = self.obj_config_software.get_log_product()

        # Nếu bật log Excel mà chưa có file -> tạo mới
        if open_log_excel and not self.path_file_save_log_excell:
            debug_print("🟢 Bật lại log Excel, tạo file mới...")
            self.path_file_save_log_excell = self.create_file_excell()
            if self.path_file_save_log_excell:
                self.write_file_excel(["Thời gian", "Mã sản phẩm", "Tên sản phẩm", "Tên người thao tác", "Mã lỗi", "Ghi chú"])
        
        # Nếu tắt log Excel mà vẫn có file -> dừng ghi
        elif not open_log_excel and self.path_file_save_log_excell:
            debug_print("🔴 Tắt log Excel (realtime cập nhật)")
            self.path_file_save_log_excell = None

#==================================Hàm chạy kiểm thử====================================================#

# from config_software import OilDetectionSystem
# obj_config_software = OilDetectionSystem()
# test_obj_log_excell = log_excell(obj_config_software)
# print("danh sach folder hien co:",test_obj_log_excell.get_list_folder_log_excell())
#print("Đường dẫn File Excell có nếu Bật log Excell",test_obj_log_excell.get_path_file_save_log_excell())
#test_obj_log_excell.get_time()
#print("Cho phéo tạo file không ?",test_obj_log_excell.get_open_log_excell())
# print("Dường dẫn lưu File excelc",test_obj_log_excell.get_path_folder_log_excell())
# test_obj_log_excell.get_list_find_old(1)
# test_obj_log_excell.delete_file_old()
# test_obj_log_excell.write_file_excel([1,3,4,5,23])
# test_obj_log_excell.write_file_excel([1,3,4,5,23])
# test_obj_log_excell.write_file_excel([1,3,4,5,23])






class log_img:
    #Kiểm thử hảm Log img Ok không cần kiểm tra lại
    from folder_create import Create
    obj_folder = Create()
    characters_check  = "date_"
    extension ="jpg"
    def __init__(self,obj_config_software):
        self.obj_config_software = obj_config_software
        self.enabled = self.get_open_log_img()  # 🔹 Khởi tạo trạng thái hiện tại
        debug_print("-------Tiến hành xóa File log ảnh-----")
        self.delete_file_old_log_img()  #xoa truoc moi khi mo phan mem
        debug_print("-----------Hoàn thành xóa-----------")
    def get_path_foldef_log_img(self):
        """Tra ve path File luu log hien tai"""
        return self.obj_config_software.get_path_log_img_oil()
    def get_time_log_img(self):
        """Lấy thời gian cho phép log được lưu nếu được bật"""
        return self.obj_config_software.GetTimeSaveLogImg()
    def get_open_log_img(self):
        """lấy quyền lưu log"""
        return self.obj_config_software.get_log_img_oil()
    def create_file_log_img(self,img):
        """Hàm này tạo lưu ảnh img khi yêu cầu bật ảnh được bật"""
        if self.get_open_log_img():
            debug_print("da vao day nha 2")
            path_foler_img = self.get_path_foldef_log_img()
            today = datetime.now().strftime("%Y-%m-%d")
            name_folder = f"date_{today}"
            path_foder = os.path.join(path_foler_img,name_folder)
            log_img.obj_folder.create_file_log_img(img,path_foder,extension= log_img.extension)
    def get_list_find_old_img(self):
            """Trả về danh sách sau khoảng thời gian time trong cấu hình information"""
            time_set  = self.get_time_log_img()
            list_file =  self.get_list_file_in_folder_img()
            return log_img.obj_folder.get_old_folders_by_threshold(log_img.characters_check,list_file,time_set)
    def delete_file_old_log_img(self):
        """Tự lấy danh sách ảnh cũ trong đường link ảnh và tự động xóa ảnh sau thời gian quá hạn"""
        arr_file_old_img = self.get_list_find_old_img()
        if arr_file_old_img:
            for file_delete in arr_file_old_img:
                path_file_delete = Log.obj_folder.find_file(self.get_path_foldef_log_img(),file_delete)
                if path_file_delete:
                    log_img.obj_folder.delete_folder(path_file_delete)
            debug_print("-------Xóa thành công folder-------")
        debug_print("Hiện tại không tìm thấy File quá hạn")
    def get_list_file_in_folder_img(self):
        return log_img.obj_folder.get_list_folder_in_folder(self.get_path_foldef_log_img())
    def update_log_state(self):
        """Kiểm tra và cập nhật trạng thái log ảnh theo cấu hình."""
        new_state = self.get_open_log_img()
        if new_state != self.enabled:
            if new_state:
                debug_print("🟢 Bật log ảnh (cho phép lưu ảnh).")
            else:
                debug_print("🔴 Tắt log ảnh (ngừng lưu ảnh).")
            self.enabled = new_state
            
#==================================Hàm chạy kiểm thử====================================================#

    
# from config_software import OilDetectionSystem
# obj_config_software = OilDetectionSystem()
# obj_log_img = log_img(obj_config_software)

# print("Thời gian lưu log hình ảnh hiện tại",obj_log_img.get_time_log_img(),"ngày")
# print("Đường dẫn lưu ảnh:",obj_log_img.get_path_foldef_log_img())
# print("Danh sách ảnh cũ quá hạn",obj_log_img.get_list_find_old_img())
# print("Trạng thái lưu log hình ảnh hiện tại",obj_log_img.get_open_log_img())
# print("Danh sách Folder ảnh cũ",obj_log_img.get_list_find_old_img())

# print("Danh sách ảnh cũ quá hạn",obj_log_img.get_list_file_in_folder_img())
# import numpy as np
# import os
# height, width, channels = 480, 640, 3
# blank_image = np.zeros((height, width, channels), dtype=np.uint8)
# obj_log_img.create_file_log_img(blank_image)
# obj_log_img.delete_file_old_log_img()

class Manager_Log:
    def __init__(self,obj_config_software,queue_log):
        self.obj_log_excell = log_excell(obj_config_software)
        self.obj_log_img  = log_img(obj_config_software)
        self.obj_log  = Log(obj_config_software)
        
        self.queue_log = queue_log
        self.thread_running = False
        self.thread = None
        self.Init()
    def Init(self):
        status_log_excell = self.obj_log_excell.get_open_log_excell()
        status_log = self.obj_log.get_open_log_software()
        status_log_cosole = self.obj_log.get_open_log_console()
        status_log_img = self.obj_log_img.get_open_log_img()
        # print(status_log_excell,status_log,status_log_img,satus_log_cosole)
        if any([
             status_log_excell,
             status_log,
             status_log_cosole,
             status_log_img 
        ]):
            self.start_log_thread()
        else:
            self.stop_log_thread()            
    def start_log_thread(self):
        """
        Khởi động luồng đọc log từ queue_log.
        Mỗi phần tử trong queue sẽ được xử lý tùy loại log.
        """
        if not self.thread_running:
            self.clear_log_queue()#xoa queue du lieu truoc khi mo luong de lam rong luong truoc
            self.thread_running = True
            self.thread = threading.Thread(target=self._log_thread_loop, daemon=True)
            self.thread.start()
     
    def stop_log_thread(self):
        """
        Dừng luồng ghi log an toàn.
        """
        self.thread_running = False
        if self.thread:
            self.thread.join(timeout=2)
           

    def _log_thread_loop(self):
        """
        Hàm chạy trong thread để đọc queue_log liên tục.
        """
        while self.thread_running:
                # Lấy dữ liệu từ queue, timeout tránh treo vô hạn
                if not self.queue_log.empty():
                    item = self.queue_log.get(timeout=0.1)
                    self._handle_log_item(item)
                time.sleep(0.01)
    def _handle_log_item(self, item):
        """
        Xử lý 1 phần tử log lấy ra từ queue.
        item có thể là dict chứa thông tin loại log, dữ liệu, v.v.
        """
        try:
            log_type = item.get("type","")
            data = item.get("data","")
            
            if log_type == "excel":
                self.obj_log_excell.write_file_excel(data)
            elif log_type == "image":
                debug_print("da vao day nha1")
                self.obj_log_img.create_file_log_img(data)
            elif log_type == "software":
                level = item.get("level",None)
                if level:
                    if level == "debug":
                            self.obj_log.logger.debug(data)
                    elif level == "warning":
                            self.obj_log.logger.warning(data)
                    elif level == "error":
                             self.obj_log.logger.error(data)
                    elif level == "critical":
                             self.obj_log.logger.critical(data)
                    elif level == "info":
                            self.obj_log.logger.info(data)
                    else:
                        debug_print("Level Hiện tại không đúng định dạng")
                else:
                    debug_print("Log Excell hiện tại chưa có level")
            else:
                self.obj_log.logger.warning(f"⚠️ Unknown log type: {log_type}")
        except Exception as e:
            self.obj_log.logger.error(f"❌ Error handling log item: {e}")
    def update_log(self):
        """
        Cập nhật trạng thái bật/tắt log.
        Nếu tất cả đều tắt → dừng luồng & xóa sạch queue_log.
        Nếu ít nhất 1 cái bật → bật luồng.
        """
        # Gọi hàm cập nhật trạng thái của từng module log
        self.obj_log.update_log_state()
        self.obj_log_img.update_log_state()
        self.obj_log_excell.update_log_state()

        # Lấy trạng thái hiện tại
        status_log_excell = self.obj_log_excell.get_open_log_excell()
        status_log = self.obj_log.get_open_log_software()
        status_log_console = self.obj_log.get_open_log_console()
        status_log_img = self.obj_log_img.get_open_log_img()

        # Nếu tất cả đều False → tắt luồng & xóa queue
        if not any([
            status_log_excell,
            status_log,
            status_log_console,
            status_log_img
        ]):
            self.stop_log_thread()
            self.clear_log_queue()
        else:
         
            self.start_log_thread()
    def clear_log_queue(self):
        """
        Xóa toàn bộ phần tử còn lại trong queue_log.
        """
        try:
            cleared = 0
            while not self.queue_log.empty():
                self.queue_log.get_nowait()
                cleared += 1
            if cleared > 0:
                debug_print(f"🧹 Đã xóa {cleared} phần tử trong queue_log.")
        except Exception as e:
            debug_print(f"⚠️ Lỗi khi xóa queue_log: {e}")

#==================================Hàm chạy kiểm thử====================================================#

# from config_software import OilDetectionSystem
# obj_config_software = OilDetectionSystem()
# import queue
# queue_data = queue.Queue(maxsize = 30)
# obj_manager_log = Manager_Log(obj_config_software,queue_data)
# obj_manager_log.start_log_thread()


    def log_and_print(self, msg, value=None, level="info"):
        # Ghép message nếu có value
        full_msg = f"{msg}: {value}" if value is not None else msg
        if level == "debug":
            self.logger.debug(full_msg)
        elif level == "warning":
            self.logger.warning(full_msg)
        elif level == "error":
            self.logger.error(full_msg)
        elif level == "critical":
            self.logger.critical(full_msg)
        else:
            self.logger.info(full_msg)

    def enable_console(self):
            debug_print("Bật Log console")
            if not self.console_enabled:
                ch = logging.StreamHandler()
                ch.setLevel(logging.DEBUG)
                ch.setFormatter(self.formatter)
                self.logger.addHandler(ch)
                self.console_enabled = True

    def disable_console(self):
            debug_print("Tắt Log console")
            for h in list(self.logger.handlers):
                if isinstance(h, logging.StreamHandler):
                    self.logger.removeHandler(h)
            self.console_enabled = False

    def enable_file(self):
            debug_print("Bật Log File")
            if not self.file_enabled:
                os.makedirs(os.path.dirname(self.log_file) or ".", exist_ok=True)
                debug_print("Đường dẫn file log:", self.log_file)
                fh = logging.FileHandler(self.log_file, encoding="utf-8")
                fh.setLevel(logging.DEBUG)
                fh.setFormatter(self.formatter)
                self.logger.addHandler(fh)
                self.file_enabled = True

    def disable_file(self):
            debug_print("Tắt Log File")
            for h in list(self.logger.handlers):
                if isinstance(h, logging.FileHandler):
                    self.logger.removeHandler(h)
            self.file_enabled = False

    # ===============================
    # Các hàm log tiện dụng
    # ===============================
    def get_open_log_excell(self):
        return self.obj_log_excell.get_open_log_excell()

    def debug(self, msg):
        self.logger.debug(msg)


    def info(self, msg):
        self.logger.info(msg)

    def warning(self, msg):
        self.logger.warning(msg)

    def error(self, msg):
        self.logger.error(msg)

    def critical(self, msg):
        self.logger.critical(msg)
    def update_log_state(self):
        """Kiểm tra và cập nhật trạng thái log theo obj_config_software (real-time)."""
        if not self.obj_config_software:
            return
        
       